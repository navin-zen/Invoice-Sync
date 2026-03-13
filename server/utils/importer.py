import csv
import logging
from io import BytesIO
from zipfile import BadZipfile

import openpyxl
import six

# from cz_utils.pusher import trigger_progress
# from cz_utils.uuid_utils import validate_uuid
# from cz_utils.import_utils import import_and_get_task
# from cz_utils.pusher import trigger_notification
# from gstnapi.utils.task_utils import scheduled_task
from openpyxl.styles import is_date_format
from tablib.core import Dataset as TablibDataset, Row

# from taxpayer.models import (
#   CachedData, GstIn, LegalPerson, TaxReturn,
# )
# from customers.models import Customer
# from cz_permissions.models import MIN_WRITER
# from cz_permissions.utils import is_admin_user
# from cz_utils.crud_views_generator import CrudViewsGenerator
# from cz_utils.dateparse import parse_date
from cz_utils.exceptions import ExcelImportException

# from cz_utils.itertools_utils import batch_list
from cz_utils.text_utils import squeeze_space

# from taxpayer.models import Attachment, ImportStatus
# from taxpayer.utils.api.common import (get_hsncode, get_or_create_legalperson, get_or_create_department, )
# from taxpayer.utils.dataimport.exception_context import ExceptionHandler

logger = logging.getLogger(__name__)


class XNum(int):
    """
    A subclass of int to use for XLS Row Numbers
    """

    pass


class RNum(int):
    """
    A subclass of int to use for dataset row number.

    This value starts from 2. Number 1 is reserved for the header of the
    dataset.

    Why not start at 0? Well, this is there for historical reasons when the
    dataset and XLS file matched with each other. Row 1 of the XLS was for
    the header. Row 2 of the XLS was for the 0th dataset row.
    """

    pass


class XlsDatasetExceptionHandler:
    """
    A context manager that handles exceptions while opening an XLS file and
    creating a dataset.

    with XlsDatasetExceptionHandler() as eh:
        ...
    """

    def __enter__(self):
        return None

    def __exit__(self, exc_type, exc_val, exc_traceback):
        if exc_type is None:
            return True
        if isinstance(exc_val, BadZipfile):
            logger.exception("Not a valid (Microsoft Excel) XLSX file.")
            raise ExcelImportException(["Not a valid (Microsoft Excel) XLSX file."])
        if isinstance(exc_val, IOError):
            if exc_val.args[0] == "File contains no valid workbook part":
                logger.exception("File contains no valid workbook part")
                raise ExcelImportException(
                    [
                        squeeze_space(
                            """File is not in a recent Microsoft XLS
                    format. Open your file, choose File -> "Save As" and
                    save in Excel Workbook format."""
                        )
                    ]
                )
            else:
                logger.exception("Got exception while creating dataset")
                raise ExcelImportException([f"Import failed because of: {exc_val}"])
        logger.exception("Got exception while creating dataset")
        raise ExcelImportException([f"Import failed because of: {exc_val}"])


class Dataset(TablibDataset):
    def insert(self, index, row, tags=list()):
        """Inserts a row to the :class:`Dataset` at the given index.

        Rows inserted must be the correct size (height or width).

        The default behaviour is to insert the given row to the :class:`Dataset`
        object at the given index.
        """

        self._validate(row, safety=True)
        self._data.insert(index, Row(row, tags=tags))


def cell_value(cell):
    try:
        return cell.value
    except ValueError:
        if (cell.data_type == "n") and (cell.style_array) and is_date_format(cell.number_format):
            # Most likely a date format error
            return cell._value
        else:
            raise
    except IndexError:
        # Happens unpredictably during testing. Not sure what is causing
        # it. Caused when accessing cell.style_array
        return cell._value


class CustomXls:
    """
    Our custom XLS format that can create dataset from XLS input.
    """

    def __init__(self, invoice_number_column_idx, *args, **kwargs):
        """
        :param: invoice_number_column_idx - The index of the column
        containing the Invoice number
        """
        # There is no super class constructor
        # super(CustomXls, self).__init__(self, *args, **kwargs)
        self.invoice_number_column_idx = invoice_number_column_idx

    @classmethod
    def _is_header_row(cls, row_values):
        """
        Try to get guess whether a row is a header row

        1) All columns are either strings or blank
        2) Has at least 5 non-blank columns
        """
        if len(row_values) < 5:
            return False
        allowed_types = (str,) + (type(None),)
        if any((not isinstance(c, allowed_types)) for c in row_values):
            return False
        if sum(isinstance(c, str) for c in row_values) < 5:
            return False
        return True

    @classmethod
    def get_header_row_from_dataset(cls, in_stream, **kwargs):
        sheet_name = kwargs.get("sheet_name", None)
        sheet_number = kwargs.get("sheet_number", None)
        sheet = cls.get_sheet_from_stream(in_stream, sheet_name, sheet_number)
        for row in sheet.rows:
            row_vals = [c.value for c in row]
            if cls._is_header_row(row_vals):
                # print(row_vals[:200])
                return row_vals[:200]
        raise ValueError("Could not find header row in the XLS file")

    @classmethod
    def get_header_row_from_dataset_wrapper(cls, xls_stream, **kwargs):
        # print(type(xls_stream))
        """
        Get the header row from the Dataset.

        Returns a list of string or None objects.

        Handles and raises exceptions
        """
        with XlsDatasetExceptionHandler():
            return cls.get_header_row_from_dataset(xls_stream, **kwargs)

    @classmethod
    def get_worksheets(cls, content):
        # with XlsDatasetExceptionHandler():
        #    return cls.get_worksheets(content)
        # if the comment on the above lines is removed, function goes into infinite loop

        xlsx_book = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        # print(xlsx_book)
        print(xlsx_book.sheetnames)
        return xlsx_book.sheetnames

        # xlsx_book.worksheets gives a worksheet object whereas xlsx_book.sheetnames gives the sheetnames
        # return xlsx_book
        # changed the return statement inorder to use xlsx_book variable in get_header_rows function

    # author: srivatsa
    # purpose: duplicate function to print headers from a worksheet
    @classmethod
    def get_header_rows(cls, content):
        xlsx_book = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        # worksheets_list = cls.get_worksheets(content)
        print(xlsx_book.sheetnames)
        sheet_obj = xlsx_book.active
        max_col = sheet_obj.max_column
        print("sheet object:", sheet_obj)
        print("max column:", max_col)
        header_list = []
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=1, column=i)
            header_list.append(cell_obj.value)
        """
        for worksheet_header in xlsx_book.sheetnames:
            sheet = worksheets_list[worksheet_header]
            for value in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                for header in value:
                    print(header)
        """
        # print(header_list)
        return [dict(r) for r in header_list]

    # author: srivatsa
    # purpose: duplicate function to print excel sheet contents
    @classmethod
    def get_excel_first_row(cls, content):
        xlsx_book = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        excel_example_row = []
        sheet = xlsx_book["Sheet2"]
        for row in sheet.rows:
            excel_example_row.append(row[0].value)
        return excel_example_row

    @classmethod
    def get_worksheets_wrapper(cls, xls_stream):
        """
        Return the list of worksheets in the XLS file.
        """
        with XlsDatasetExceptionHandler():
            return cls.get_worksheets(xls_stream.read())

    @classmethod
    def get_sheet_from_stream(cls, xls_stream, sheet_name, sheet_number):
        """
        Read the Workbook from `xls_stream` and then get the Sheet.
        Review this code later
        """
        # gc = GlobalConfiguration.get_solo()
        # path = gc.metadata['datasource']['config']['path']
        xlsx_book = openpyxl.load_workbook(xls_stream)
        return cls.get_sheet(xlsx_book, sheet_name, sheet_number)

    @classmethod
    def get_sheet(cls, xlsx_book, sheet_name, sheet_number):
        """
        Get the Sheet from the XLSX file.

        :param: xlsx_book - The XLS Workbook
        :param: sheet_name - The name of the sheet to use (None indicates don't use sheet_name)
        :param: sheet_number - The number of the sheet (-1 or None indicates don't use sheet_number)

        We first try the sheet_number, otherwise the sheet_name, else the
        active sheet.
        """
        assert (sheet_name is None) or isinstance(sheet_name, str)
        assert (sheet_number is None) or isinstance(sheet_number, int)
        if (sheet_number is not None) and (sheet_number >= 0):
            try:
                return xlsx_book.worksheets[sheet_number]
            except IndexError:
                raise ValueError(
                    squeeze_space(
                        """The XLSX file does not contain the
                    sheet number {}. Please check the uploaded file and try
                    again."""
                    ).format(sheet_number + 1)
                )
        elif sheet_name:
            try:
                return xlsx_book[sheet_name]
            except KeyError:
                raise ValueError(
                    squeeze_space(
                        """The XLSX file does not contain the
                    sheet {}. Please check the uploaded file and try
                    again."""
                    ).format(sheet_name)
                )
        else:
            return xlsx_book.active

    @classmethod
    def create_dataset(cls, in_stream, **kwargs):
        """
        Create dataset for QuickBooks xls format

        Returns (dataset, row_num_mapping)

        dataset - a tablib.Dataset object
        row_num_mapping - Mapping from
            dataset row number (2 based, the header is row number 1) -> XLS row number (1 based)

        NOTE: xls_rows_range - The range of rows to consider while
        importing the dataset. The range is inclusive.
        """
        sheet_name = kwargs.get("sheet_name", None)
        sheet_number = kwargs.get("sheet_number", None)
        xls_rows_range = kwargs.get("xls_rows_range", None)
        invoice_number_column_idx = kwargs.get("invoice_number_column_idx", None)
        if xls_rows_range:
            (first_row, last_row) = xls_rows_range
        else:
            (first_row, last_row) = (0, six.MAXSIZE)
        # if (self.invoice_number_column_idx is None):
        # raise ValueError("Index of Invoice Number Column is not specified")

        sheet = cls.get_sheet_from_stream(in_stream, sheet_name, sheet_number)
        found_header = False
        dset = Dataset()
        row_num_mapping = {}
        for i, row in enumerate(sheet.rows, start=1):
            row_vals = [cell_value(c) for c in row]
            if not found_header:
                if not cls._is_header_row(row_vals):
                    continue
                dset.headers = row_vals
                found_header = True
                row_num_mapping[RNum(1)] = XNum(i)
            else:
                if not (first_row <= i <= last_row):
                    continue
                try:
                    if invoice_number_column_idx is not None:
                        invoice_number = row_vals[invoice_number_column_idx]
                        # Only consider rows that have an Invoice number
                        #
                        # Handle the corner case where the Invoice number could
                        # be the number 0 or the float 0.0. Such rows are valid
                        # Invoices.
                        #
                        # Rows will all spaces. We should exclude those rows
                        if (
                            (not isinstance(invoice_number, int))
                            and (not isinstance(invoice_number, float))
                            and (
                                (not invoice_number) or (isinstance(invoice_number, str) and not invoice_number.strip())
                            )
                        ):
                            continue
                except IndexError:  # The row has only a few columns
                    continue
                dset.append(row_vals)
                row_num_mapping[RNum(len(dset) + 1)] = XNum(i)

        # header_list = cls.get_header_row_from_dataset_wrapper(in_stream)
        return (dset, row_num_mapping)

    @classmethod
    def create_example_row(cls, in_stream):
        dset = cls.create_dataset(in_stream)[0]
        header_list = cls.get_header_row_from_dataset_wrapper(in_stream)
        return [dict(zip(header_list, dset[0]))]


class CustomXlsNoHeaderSearch(CustomXls):
    """
    A simplification of CustomXls above. Does not search for the header
    row. Does not know about Invoice Number column. This is useful when we
    simply want the dataset from an XLS sheet.

    Used in our tests.
    """

    def __init__(self, *args, **kwargs):
        """
        No need to call super class __init__.
        """
        pass

    def create_dataset(self, in_stream, **kwargs):
        sheet_name = kwargs.get("sheet_name", None)
        sheet_number = kwargs.get("sheet_number", None)
        sheet = self.get_sheet_from_stream(in_stream, sheet_name, sheet_number)
        dataset = Dataset()
        # obtain generator
        rows = sheet.rows
        dataset.headers = [cell.value for cell in next(rows)]
        for row in rows:
            row_values = [cell.value for cell in row]
            dataset.append(row_values)
        return dataset


class CustomCsv:
    """
    Our custom CSV format that can create dataset from CSV input.
    """

    @classmethod
    def create_dataset(cls, in_stream, **kwargs):
        dataset = Dataset()
        rows = csv.reader((line.replace("\0", "") for line in in_stream), delimiter=",")
        dataset.headers = [cell for cell in next(rows)]
        for row in rows:
            row_values = [(cell.strip() if isinstance(cell, str) else cell) for cell in row]
            dataset.append(row_values)
        return dataset

    @classmethod
    def get_header_row(cls, in_stream):
        reader = csv.reader(in_stream)
        row1 = next(reader)
        return list(row1)

    @classmethod
    def create_example_row(cls, in_stream):
        dset = cls.create_dataset(in_stream)
        return [{}]
        header_list = cls.get_header_row(in_stream)
        return [dict(zip(header_list, dset[0]))]


def get_xls_stream(attachment):
    """
    Get the file stream of the Attachment object.

    This is a separate function so that we can mock it during testing
    """
    return attachment.file_obj
